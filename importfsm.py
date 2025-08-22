#!usr/bin/env python

import os, shutil, warnings, pdfplumber, logging
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
# QUIET, QUIET ##############################################################################
warnings.simplefilter("ignore", UserWarning)                                                # Ignore specific openpyxl warnings
logging.getLogger("pdfminer").setLevel(logging.ERROR)                                       # Prevents "CropBox missing from /Page, defaulting to MediaBox" spam
blue_fill = PatternFill(start_color="00b0f0", end_color="00b0f0", fill_type="solid")        # This for conditional formatting via openpyxl - cell needs input
green_fill = PatternFill(start_color="00b050", end_color="00b050", fill_type="solid")       # This for conditional formatting via openpyxl - DUT is good to go
yellow_fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")      # This for conditional formatting via openpyxl - DUT needs something/FI
red_fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")         # This for conditional formatting via openpyxl - DUT should be recalled
# HELPER FUNCTIONS ##########################################################################
def blue_if_blank_formatting(sheet, ranges):                                                # Add conditional formatting to highlight blank cells with blue fill for the given list of range strings.
    for rng in ranges:                                                                      # Iterate over the ranges provided
        first_cell = rng.split(":")[0]                                                      # Get the first element of the cell
        formula = f'=OR(ISBLANK({first_cell}), {first_cell}="")'                            # Establish the conditional formatting formula
        rule = FormulaRule(formula=[formula], fill=blue_fill)                               # Establish the rule, for the formula above
        sheet.conditional_formatting.add(rng, rule)                                         # add the rule to the sheet provided

def TUR_check_formatting(sheet, start_row, end_row):                                        # More conditional formatting based on Column X - the preliminary evaluation
    target_range = f"A{start_row}:X{end_row}"
    col_x_first = f"$X{start_row}"
    sheet.conditional_formatting.add(
        target_range,
        FormulaRule(formula=[f'={col_x_first}="Not Significant"'], fill=green_fill)
    )
    sheet.conditional_formatting.add(
        target_range,
        FormulaRule(formula=[f'={col_x_first}="Semi-Significant"'], fill=yellow_fill)
    )
    sheet.conditional_formatting.add(
        target_range,
        FormulaRule(formula=[f'={col_x_first}="Significant"'], fill=red_fill)
    )

def final_eval_formatting(sheet, start_row, end_row):                                       # More conditional formatting based on Column AI - the final evaluation
    whole_row = f"A{start_row}:AJ{end_row}"
    col_x_first = f"$AI{start_row}"
    sheet.conditional_formatting.add(
        whole_row,
        FormulaRule(formula=[f'={col_x_first}="No Further Action Required"'], fill=green_fill)
    )
    sheet.conditional_formatting.add(
        whole_row,
        FormulaRule(formula=[f'={col_x_first}="Analysis performed; no further action required."'], fill=green_fill)
    )
    sheet.conditional_formatting.add(
        whole_row,
        FormulaRule(formula=[f'={col_x_first}="No intersection; no further action required."'], fill=green_fill)
    )
    sheet.conditional_formatting.add(
        whole_row,
        FormulaRule(formula=[f'={col_x_first}="FI Fails; TSM determination required."'], fill=yellow_fill)
    )
    sheet.conditional_formatting.add(
        whole_row,
        FormulaRule(formula=[f'={col_x_first}="Unit fails analysis; TSM determination required."'], fill=red_fill)
    )
    sheet.conditional_formatting.add(
        whole_row,
        FormulaRule(formula=[f'={col_x_first}="Significant preliminary finding & no data; TSM determination required."'], fill=red_fill)
    )

# MAIN LOOP #################################################################################
def main():                                                                                 # The main loop of the application
    # INIT ##################################################################################
    current_dir = Path.cwd()                                                                # where script was run from
    this_oot_dir = current_dir                                                              # where script was run from
    code_dir = Path(__file__).parent                                                        # /home/datavikingr/Tek/OOTs/auto-oot
    oots_dir = code_dir.parent                                                              # /home/datavikingr/Tek/OOTs
    ds_filename = "DS.pdf"                                                                  # This is the datasheet we'll import.
    rev_trace = "Reverse Trace.xlsx"                                                        # This is the source rev trace data
    tek_logo = code_dir / "Tek_logo.png"                                                    # /home/datavikingr/Tek/OOTs/auto-oot/Tek_logo.png
    oot_template = code_dir / "FSMOOTSIA.xlsm"                                              # /home/datavikingr/Tek/OOTs/auto-oot/FSMOOTSIA.xlsm - This is the template we're reproducing
    today = datetime.today().strftime('%m/%d/%Y')                                           # This goes in the header of output file
    this_year = datetime.today().strftime('%Y')                                             # This is a question for directory-location
    start_row = 10                                                                          # This is the header of the analysis itself, and is row(0), effectively
    end_row = 5008                                                                          # This is the current maximum size of these analyses. Adjust THIS number, if we need to analyze assets that have intersected with more than 5k assets
    # ERROR CHECKING ########################################################################
    rev_trace_file = this_oot_dir / rev_trace                                               # Establish the file we need
    if not rev_trace_file.exists():                                                         # This is a good proxy to see what kind of directory we're running the script in - if there isn't a rev trace in cwd, we're probably running it from $HOME, so we need to build the correct location and 'navigate' to it
        lab = input("Enter lab location: ")                                                 # Baltimore, Strother, etc
        asset_UID = input("Enter UID: ")                                                    # The asset itself
        this_oot_dir = oots_dir / f"{this_year}/{lab}/{asset_UID}/"                         # Establish the output directory
        this_oot_dir.mkdir(parents=True, exist_ok=True)                                     # Ensure output dir exists
        rev_trace_file = this_oot_dir / rev_trace                                           # And then re-set to where the rev trace must be.
    else:                                                                                   # Then we have a rev trace file, and we need to set this one variable that's required later
        asset_UID = this_oot_dir.name                                                       # Because I'm not bloody asking the techs to type UIDs if I can avoid it
    if not oot_template.exists():                                                           # Let's check for the template file, see if the package was tampered with
        file = oot_template.name                                                            # Well, it's gone. So, let's get the name out of the file path,
        print(f"{file} does not exist.")                                                    # so we can announce what's happened
        raise SystemExit(1)                                                                 # ...and gtfo
    ds_import = input("Import datasheet? ").strip().lower()                                 # Final bit of INIT input for later. TODO: set this up in the GUI, when that time comes
    # LOAD WORKBOOK, SHEETS ################################################################# HACK I just learned that opnepyxl plays nice with path objects, so I don't have to refactor most of this as I rebuild around pathlib. Very excited!! - AJH 21AUG25
    oot_wb = load_workbook(oot_template, data_only=False, keep_vba=False)                   # Load the workbook, so we can get its sheets (read: tabs) 
    oot_rt = oot_wb["Reverse Trace"]                                                        # FSMOOTSIA.xlsm > Reverse Trace (tab); receives FSM's rev trace data
    oot_ia = oot_wb["Impact Analysis"]                                                      # FSMOOTSIA.xlsm > Impact Analysis (tab); where we're doing the dirty work.
    rt_wb = load_workbook(rev_trace_file)                                                   # Load the FSM-supplied Reverse Trace Workbook
    rt_ws = rt_wb["Reverse Trace - UID"]                                                    # Load the correct sheet from the rt_wb, so we can get at the data
    # HEADER DATA ###########################################################################
    oot_uid =  rt_ws['D2'].value                                                            # Get the Asset's UID
    oot_ia['D1'] = oot_uid                                                                  # Assign it to the header of the template
    owning_lab = rt_ws['F2'].value                                                          # Get the owning lab
    oot_ia['D2'] = owning_lab                                                               # Assign it to the header of the template
    prev_cal = rt_ws['G2'].value                                                            # HACK Get the previous calibration date
    oot_ia['D3'] = prev_cal                                                                 # Assign it to the header of the template
    curr_cal = rt_ws['H2'].value                                                            # HACK Get the current calibration date, completing the date range for the cal cycle
    oot_ia['D4'] = curr_cal                                                                 # Assign it to the header of the template
    analysis_date = today                                                                   # When was the analysis performed? Today, duh.
    oot_ia['D5'] = analysis_date                                                            # Assign it to the header of the template
    last_row = rt_ws.max_row - 1                                                            # Get how many assets there are
    oot_ia['H1'] = last_row                                                                 # Assign it to the header of the template
    # COPY RAW REVERSE TRACE DATA INTO TEMPLATE WB ##########################################
    for row in rt_ws.iter_rows():                                                           # Each row in rev trace raw data
        for cell in row:                                                                    # Each cell in that data
            new_cell = oot_rt.cell(row=cell.row, column=cell.column, value=cell.value)      # Copy it into the template's Reverse Trace tab
    rt_wb.close()                                                                           # Close the Reverse Trace workbook, we no longer need it as we got the data we came for already
    # TAKE RECENTLY IMPORTED DATA AND MOVE IT TO ANALYSIS LOCATION ##########################
    col_map = [
        ('J', 'A'), ('K', 'B'), ('L', 'C'),
        ('O', 'D'), ('Q', 'F'), ('R', 'G'),
        ('S', 'I'), ('M', 'J'), ('N', 'K'),
        ('P', 'L')]                                                                         # Map the columns to each other
    for src, dst in col_map:                                                                # Iterate over the mapped pairs
        values = [oot_rt[f"{src}{row}"].value for row in range(2, last_row + 2)]            # Get the data in the filled in the source range
        for i, value in enumerate(values, start=10):                                        # Get the spaces in destination range
            oot_ia[f"{dst}{i}"].value = value                                               # Dump the source data into the destination ranges
    # SAMPLE METHOD INIT  ###################################################################
    col = 'A'                                                                               # We should examine Col A for blanks, so we can hide them
    last_row_in_range = start_row                                                           # Start from the beginning at "Analysis Row 0"
    sort_column = 2                                                                         # Column B, in number-count, so we can check Products, as our sample-method relies on unique 'products' which mfgr/model pairs
    for row in range(start_row, end_row + 1):                                               # Get the actual range of the sheet, as well as hides empty-product fields
        if oot_ia[f"{col}{row}"].value is not None:                                         # Check each cell in Col A for blanks
            last_row_in_range = row                                                         # if not blank, we set the new last row, and continue the loop. At the end, we find the true last row.
        else:                                                                               # if blank, however, we've found the end and the latest last_row_in_range continues the very last asset's row
            oot_ia.row_dimensions[row].hidden = True                                        # we can just hide this row now, and continue the loop, instead of iterating over the last_row_in_range to end_row (5008) in a second loop.
    # EXTRACT ALL ROWS OF DATA FOR SORT #####################################################
    table_data = []                                                                         # Establish a clean data object
    for row in range(start_row, last_row_in_range + 1):                                     # Iterate over the actual range of the data
        row_data = [oot_ia.cell(row=row, column=col).value for col in range(1, 13)]         # Constrain the column data to A-M - prevents other issues later
        table_data.append(row_data)                                                         # Stuff it into the Table object
    table_data.sort(key=lambda x: x[sort_column - 1])                                       # Sort the rows in the data table based on Column B, adjusting for 0-based index
    for idx, row_data in enumerate(table_data, start=start_row):                            # Iterate over the rows of table_data
        for col_idx, value in enumerate(row_data, start=1):                                 # Grab each cell, so we can...
            oot_ia.cell(row=idx, column=col_idx).value = value                              # ...distribute the sorted data back into the worksheet
    # HIDE DUPLICATE PRODUCTS ############################################################### NOTE This is the actual sample occuring, every leading up to now has been prep for it
    for product in range(start_row + 1, last_row_in_range + 1):                             # Iterate over each row to compare product values in column B
        if oot_ia[f"B{product}"].value == oot_ia[f"B{product-1}"].value:                    # if this row's product matches the previous product, then...
            oot_ia.row_dimensions[product].hidden = True                                    # ...hide the entire row
    # CLEAN UP FROM THE DATA IMPORT #########################################################
    oot_ia['M10'] = ""                                                                      # So. The form's first actual data row, row 11, needs to be blank. This is because most the sheet will refer to this data, and duplicate it. 
    oot_ia['N10'] = ""                                                                      # Why duplicate it, instead of listing it once and being done forever? That is because not all rows will ACTUALLY need those particular data
    oot_ia['O10'] = ""                                                                      # And those particular data are just the starting point for the analysis itself. 
    oot_ia['P10'] = ""                                                                      # So we're building the jumping off point, for this entire imapct analysis sheet
    oot_ia['C10'].number_format = 'MM/DD/YYYY'                                              # And column C, for some reason, simply refuses to stay Date-formatted.
    for row in range(11, last_row_in_range + 1):                                            # And then, from there, we establish the referencing mentioned above, by iterating over Rows 12-last_row
        oot_ia[f"C{row}"].number_format = 'MM/DD/YYYY'                                      # And FORCING DATE FORMAT
        oot_ia[f'M{row}'] = '=IF($M$10="","",$M$10)'                                        # And then referencing the top-most row. This way, if we need to compare other test points (not uncommon)
        oot_ia[f'N{row}'] = '=IF($N$10="","",$N$10)'                                        # We can just overwrite these data and input the new data
        oot_ia[f'O{row}'] = '=IF($O$10="","",$O$10)'                                        # and get the same level of analysis and comparison to ensure our customers' equipment
        oot_ia[f'P{row}'] = '=IF($P$10="","",$P$10)'                                        # was or was not affected by the OOT Condition of the errant standard
        if oot_ia[f'A{row}'].value is None:                                                 # And finally, we double check that Column A's data isn't blank
            oot_ia.row_dimensions[row].hidden = True                                        # Because if it is, we hide it.
    # DATASHEET IMPORT ###################################################################### NOTE: Only works on Tek Datasheets
    if ds_import in ["y", "yes"]:                                                           # Remember asking this in the last line of error checking?
        ds_file = this_oot_dir / ds_filename                                                # Establish the datasheet file
        if not ds_file.exists():                                                            # pretty clear: if it doesn't exist, then...
            file = ds_file.name                                                             # establish just the file name
            print(f"{file} does not exist.")                                                # report the problem and move on
        else:                                                                               # But if it does exist
            with pdfplumber.open(ds_file) as pdf:                                           # we crack that bad boy open with pdfplumber
                all_table_data = []                                                         # Establish as clean data_table for this purpose
                for page in pdf.pages:                                                      # Iterate over every page in the pdf
                    table = page.extract_table()                                            # pdfplumber.extract_table() pulls the data and best guesses at an excel-like table format; it's pretty okay
                    if table:                                                               # if that succeeds and grabs data,
                        all_table_data.extend(table)                                        # then we pop this page's data into the formerly-clean table object established above
                tableData = all_table_data                                                  # Duplicate the object, so we can interact with that
            if tableData:                                                                   # Make sure that processed okay
                # EXTRACT ONLY DATA BETWEEN KEYWORDS ########################################
                start_keyword = "Function"                                                  # This is keyword 1, found immediately prior to the relevant data
                end_keyword = "Decision Rule"                                               # This is keyword 2, found immediately after the relevant data
                filtered_table = []                                                         # yet another clean tablwe object
                recording = False                                                           # This is the flag to see if we're pulling that row of data
                for row in tableData:                                                       # for each row of data:
                    if row and any(start_keyword in str(cell) for cell in row if cell):     # see if it's after the start keyword
                        recording = True                                                    # if it is, we're recording it
                    if recording:                                                           # if we're recording it,
                        filtered_table.append(row)                                          # we'll add it to our filtered dataset
                    if row and any(end_keyword in str(cell) for cell in row if cell):       # see if it's after the end keyword
                        break                                                               # and remove that
                # REMOVE LAST ROW - EXTRANEOUS CATCH ########################################
                if filtered_table and any(end_keyword in str(cell) for cell in filtered_table[-1] if cell): # containing "Decision Rule"
                    filtered_table.pop()                                                    # kick that mother- outta there
                # REGENERARTE THE DATASHEET IN EXCEL ########################################
                if "Datasheet" in oot_wb.sheetnames:                                        # This is just a clean up subroutine
                    ws_old = oot_wb["Datasheet"]                                            # find any old Datasheet tab
                    oot_wb.remove(ws_old)                                                   # kill it
                oot_ds = oot_wb.create_sheet("Datasheet", index=0)                          # Then we're going to add a new one
                for row in filtered_table:                                                  # And for each row in the filtered DS data
                    oot_ds.append(row)                                                      # Dump it into the Datasheet tab in Excel
            else:                                                                           # if we couldn't find the table data after pulling it out of the pdf
                print("Could not extract data.")                                            # then we report the issue and move on
    # DUPLICATE IMPACT ANALYSIS TAB FOR EACH OOT PARAMETER ##################################
    parameter_list = {}                                                                     # Start with a clean parameter list
    for idx, row in enumerate(oot_ds.iter_rows(min_row=1, values_only=True), start=1):      # iterate over the rows in the ds tab 
        col_a = row[0]                                                                      # establish Col A in this paradigm
        col_b = row[1] if len(row) > 1 else None                                            # establish Col B (or None) - and this is our checker for a "parameter" row
        if col_a and not col_b:                                                             # IF Column A has text AND Column B is empty
            parameter_list[str(col_a)] = idx                                                # key = Column A text, value = row number; now we have a list of the grey parameter bars from the Tek DS
    # FIND FAILURES IN COL D ################################################################
    failures = []                                                                           # Clean failures list
    for idx, row in enumerate(oot_ds.iter_rows(min_row=1, values_only=True), start=1):      # iterate over the DS tab's rows
        col_d = row[3] if len(row) > 3 else None                                            # Column D is index 3
        if col_d == "Fail":                                                                 # IF matches exactly "Fail", not "Fail*"
            failures.append(idx)                                                            # THEN add it to the list
    fail_formula = '=$D1="Fail"'                                                            # Establish a conditional formatting rule
    oot_ds.conditional_formatting.add(f"D1:D{oot_ds.max_row}",FormulaRule(formula=[fail_formula],fill=red_fill)) # Add it the DS Tab, so we can quickly see failures throughout the DS
    sorted_params = sorted(parameter_list.items(), key=lambda x: x[1])                      # Sort parameters by row number
    param_ranges = {}                                                                       # Clean dict of param_name: (param_start_row, param_end_row)
    for i, (param, param_start_row) in enumerate(sorted_params):                            # iterate over the sorted parameters
        if i + 1 < len(sorted_params):                                                      # if next param is on row 90
            param_end_row = sorted_params[i + 1][1] - 1                                     # then this param's last row is 89
        else:                                                                               # OR
            param_end_row = oot_ds.max_row                                                  # we've hit the end of the list, and that should be this param's last row #
        param_ranges[param] = (param_start_row, param_end_row)                              # now build the dictioary of each parameters' ranges
    # MATCH FAILURES TO PARAMETERS ########################################################## NOTE We're matching fails to parameters
    oot_parameters = []                                                                     # Clean parameters list
    for fail_row in failures:                                                               # for each failed test point
        for param, (start, end) in param_ranges.items():                                    # iterate over the dictioary of params and their ranges
            if start <= fail_row <= end:                                                    # if this fail between the start/end of X param
                oot_parameters.append(param)                                                # add it to the list of oot parameters
                break                                                                       # and then break out of this sub-loop, to move on to the failed test point
    oot_parameters = list(dict.fromkeys(oot_parameters))                                    # And then we drop everything but the parameter names, for a single list of parameters that require analysis
    # BUILD NEW ANALYSIS SHEETS PER FAILED PARAMETER ######################################## NOTE and those new sheets need conditional formatting!
    target_ranges = [f"M{start_row}:P{end_row}", 
        f"R{start_row}:S{end_row}", 
        f"Y{start_row}:Z{end_row}"]                                                         # Establish our conditional formatting ranges
    for param in oot_parameters:                                                            # Iterate over the list of failed parameters
        new_sheet = oot_wb.copy_worksheet(oot_ia)                                           # Build that new sheet per parameter
        new_sheet.title = param                                                             # Rename the sheet to the OOT parameter
        new_sheet.freeze_panes = "B10"                                                      # We always want to see the table headers and UID column
        dv = DataValidation(type="list", formula1="=$AD$1:$AD$6", allow_blank=True)         # Establish the data validation 'rule'
        new_sheet.add_data_validation(dv)                                                   # make it exist in the worksheet
        dv.add(f"$AI$10:$AI${last_row_in_range}")                                           # and then attach it to the correct cells
        final_eval_formatting(new_sheet, start_row, end_row)                                # We have to add the conditional formatting
        TUR_check_formatting(new_sheet, start_row, end_row)                                 # in order of their precedence, to achieve 
        blue_if_blank_formatting(new_sheet, target_ranges)                                  # the desired formatting throughout the analysis
        if tek_logo.exists():                                                               # Check to see the logo file exists
            logo_img = XLImage(tek_logo)                                                    # Go ahead and grab that
            new_sheet.add_image(logo_img, "A1")                                             # And slap it into the new sheets
    # FINAL CLEANUP #########################################################################
    oot_wb.remove(oot_ia)                                                                   # remove the extraneous template sheet
    oot_out_file = f"OOT_{asset_UID}.xlsx"                                                  # Build the filename
    final_xlsx = this_oot_dir / oot_out_file                                                # Build the new filepath, as a path object
    oot_wb.save(final_xlsx)                                                                 # Save the new file
    oot_wb.close()                                                                          # Close the workbook entirely

# MAIN LOOP INITIATION ######################################################################
if __name__ == "__main__":                                                                  # If we called this directly, and NOT if we loaded this as a library
    main()                                                                                  # "Zhu Li, do the thing!"