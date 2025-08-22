#!/usr/bin/env python3

import os, shutil, warnings, pdfplumber
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Optional pandas use (kept safe if not installed)
try:
    import pandas as pd
    HAVE_PANDAS = True
except Exception:
    HAVE_PANDAS = False

# Ignore specific openpyxl warnings
warnings.simplefilter("ignore", UserWarning)

current_dir = os.getcwd()
oot_dir = current_dir
today = datetime.today().strftime('%m/%d/%Y')
this_year = datetime.today().strftime('%Y')

rev_trace = "Reverse Trace.xlsx"
oot_template = "/home/datavikingr/Tek/OOTs/FSMOOTSIA.xlsm"
oot_temp_file = "FSMOOTSIA.xlsm"   # name of the copied template (xlsm), used only as a source
oot_out_file = "FSMOOTSIA_test.xlsx"    # final usable Excel file (xlsx)
ds_filename = "DS.pdf"

# ---------------------------
# Helpers
# ---------------------------
def norm_text(x):
    if x is None:
        return ""
    return " ".join(str(x).split()).strip()

def is_grey_rgb(rgb):
    """
    pdfplumber gives non_stroking_color as [r,g,b] floats 0..1 (sometimes None).
    Treat 'grey' as near-equal channels and not too close to 0 or 1.
    """
    if not rgb or not isinstance(rgb, (list, tuple)) or len(rgb) < 3:
        return False
    r, g, b = rgb[:3]
    if any(v is None for v in (r, g, b)):
        return False
    # nearly equal channels
    if abs(r - g) < 0.05 and abs(r - b) < 0.05 and abs(g - b) < 0.05:
        # not almost white or almost black
        return 0.2 <= r <= 0.9
    return False

def extract_parameter_headers_from_pdf(pdf_path):
    """
    Find grey, wide rectangles on each page and pull the text inside each as a header.
    Returns a list of header strings, top-to-bottom, across pages.
    """
    headers = []
    if not os.path.exists(pdf_path):
        return headers

    with pdfplumber.open(pdf_path) as pdf:
        for p_idx, page in enumerate(pdf.pages):
            page_w = page.width
            page_h = page.height

            candidates = []
            # page.rects are vector rectangles on the page
            for rect in page.rects:
                x0, y0, x1, y1 = rect.get("x0"), rect.get("y0"), rect.get("x1"), rect.get("y1")
                w = (x1 - x0) if (x1 is not None and x0 is not None) else 0
                h = (y1 - y0) if (y1 is not None and y0 is not None) else 0

                nonstroke = rect.get("non_stroking_color")  # fill color
                # Heuristics: very wide, short-ish, grey fill
                if w >= 0.80 * page_w and 8 <= h <= 60 and is_grey_rgb(nonstroke):
                    # small inset to avoid borders
                    pad = 1.5
                    bbox = (x0 + pad, y0 + pad, x1 - pad, y1 - pad)
                    region = page.within_bbox(bbox)
                    text = norm_text(region.extract_text())
                    if text:
                        # y1 is the top edge (larger y is higher on page)
                        candidates.append((p_idx, y1, text))

            # Top-to-bottom within page (y1 high → top)
            candidates.sort(key=lambda t: -t[1])
            headers.extend([t[2] for t in candidates])

    # De-duplicate while preserving order (PDFs sometimes draw overlaying rects)
    seen = set()
    ordered = []
    for h in headers:
        if h not in seen:
            seen.add(h)
            ordered.append(h)
    return ordered

def sheet_find_first_row_for_header(ws, start_row, header_text):
    """
    Find the first row index at or after start_row whose row (any cell) contains header_text (case-insensitive).
    Returns row index or None.
    """
    target = header_text.lower()
    for r in range(start_row, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_vals.append(str(v))
        if row_vals:
            joined = " | ".join(row_vals).lower()
            if target in joined:
                return r
    return None

def region_has_exact_fail_col_d(ws, row_start, row_end):
    """
    Check if any cell in column D (4) equals exactly 'Fail' (case-insensitive), ignoring spaces.
    Search rows [row_start, row_end] inclusive.
    """
    for r in range(row_start, row_end + 1):
        v = ws.cell(row=r, column=4).value  # column D
        if v is None:
            continue
        v_norm = norm_text(v).lower()
        if v_norm == "fail":  # exact match only
            return True
    return False

# ---------------------------
# Light error checking
# ---------------------------
rev_trace_file = os.path.join(current_dir, rev_trace)
if not os.path.exists(rev_trace_file):
    print(f"{rev_trace_file} does not exist.")
    lab = input("Enter lab location: ")  # Baltimore, Strother
    UID = input("Enter UID: ")
    oot_dir = f"/home/datavikingr/Tek/OOTs/{this_year}/{lab}/{UID}/"
    rev_trace_file = os.path.join(oot_dir, rev_trace)
if not os.path.exists(oot_template):
    print(f"{oot_template} does not exist.")
    raise SystemExit(1)

ds_import = input("Import datasheet? ").strip().lower()

# Ensure output dir exists
os.makedirs(oot_dir, exist_ok=True)

# Copy the template (xlsm) into the working directory so we read sheets from it
src_template_copy = os.path.join(oot_dir, oot_temp_file)
shutil.copy(oot_template, src_template_copy)

# Load the template copy (we will save to XLSX later; no rename)
oot_wb = load_workbook(src_template_copy, data_only=False, keep_vba=False)
oot_rt = oot_wb["Reverse Trace"]
oot_ia = oot_wb["Impact Analysis"]

# Load Reverse Trace file (source of data)
rt_wb = load_workbook(rev_trace_file)
rt_ws = rt_wb["Reverse Trace - UID"]

# Grab header data
oot_uid = rt_ws['D2'].value
owning_lab = rt_ws['F2'].value
prev_cal = rt_ws['G2'].value
curr_cal = rt_ws['H2'].value
analysis_date = today
last_row = rt_ws.max_row - 1

# Assign headers
oot_ia['D1'] = oot_uid
oot_ia['D2'] = owning_lab
oot_ia['D3'] = prev_cal
oot_ia['D4'] = curr_cal
oot_ia['D5'] = analysis_date
oot_ia['H1'] = last_row

# Copy Reverse Trace into template
for row in rt_ws.iter_rows():
    for cell in row:
        oot_rt.cell(row=cell.row, column=cell.column, value=cell.value)
rt_wb.close()

# Mappings from Reverse Trace → Impact Analysis
col_map = [
    ('J', 'A'), ('K', 'B'), ('L', 'C'),
    ('O', 'D'), ('Q', 'F'), ('R', 'G'),
    ('S', 'I'), ('M', 'J'), ('N', 'K'),
    ('P', 'L')
]
for src, dst in col_map:
    values = [oot_rt[f"{src}{row}"].value for row in range(2, last_row + 2)]
    for i, value in enumerate(values, start=10):
        oot_ia[f"{dst}{i}"].value = value

# Sorting / hiding same as before
col = 'A'
start_row = 10
end_row = 5008
last_row_in_range = start_row
sort_column = 2  # Column B
for row in range(start_row, end_row + 1):
    if oot_ia[f"{col}{row}"].value is not None:
        last_row_in_range = row
    else:
        oot_ia.row_dimensions[row].hidden = True

# Extract, sort, write back
table_data = []
for row in range(start_row, last_row_in_range + 1):
    row_data = [oot_ia.cell(row=row, column=c).value for c in range(1, oot_ia.max_column + 1)]
    table_data.append(row_data)
table_data.sort(key=lambda x: x[sort_column - 1])
for idx, row_data in enumerate(table_data, start=start_row):
    for col_idx, value in enumerate(row_data, start=1):
        oot_ia.cell(row=idx, column=col_idx).value = value

# Hide duplicate products
for r in range(start_row + 1, last_row_in_range + 1):
    if oot_ia[f"B{r}"].value == oot_ia[f"B{r-1}"].value:
        oot_ia.row_dimensions[r].hidden = True

# Final clean up
oot_ia['M10'] = ""
oot_ia['N10'] = ""
oot_ia['O10'] = ""
oot_ia['P10'] = ""
oot_ia['C10'].number_format = 'MM/DD/YYYY'
for row in range(11, last_row_in_range + 1):
    oot_ia[f"C{row}"].number_format = 'MM/DD/YYYY'
    oot_ia[f'M{row}'] = '=IF($M$10="","",$M$10)'
    oot_ia[f'N{row}'] = '=IF($N$10="","",$N$10)'
    oot_ia[f'O{row}'] = '=IF($O$10="","",$O$10)'
    oot_ia[f'P{row}'] = '=IF($P$10="","",$P$10)'
    if oot_ia[f'A{row}'].value is None:
        oot_ia.row_dimensions[row].hidden = True

# ============================
# PDF → Datasheet Import + Parameter detection from grey bars
# ============================
parameter_flags = {}  # ordered dict behavior via insertion order (py3.7+)
header_rows_in_sheet = {}  # map header → row index in Datasheet

if ds_import in ["y", "yes"]:
    ds_file = os.path.join(oot_dir, ds_filename)
    if not os.path.exists(ds_file):
        print(f"{ds_file} does not exist.")
    else:
        # Extract the main table body (your original logic)
        with pdfplumber.open(ds_file) as pdf:
            all_table_data = []
            for page in pdf.pages:
                table = page.extract_table()
                if table:
                    all_table_data.extend(table)
            tableData = all_table_data

        if tableData:
            # Extract only the section between "Function" and "Decision Rule"
            start_keyword = "Function"
            end_keyword = "Decision Rule"
            filtered_table = []
            recording = False
            for row in tableData:
                if row and any(start_keyword in str(cell) for cell in row if cell):
                    recording = True
                if recording:
                    filtered_table.append(row)
                if row and any(end_keyword in str(cell) for cell in row if cell):
                    break
            if filtered_table and any(end_keyword in str(cell) for cell in filtered_table[-1] if cell):
                filtered_table.pop()

            # (Re)create Datasheet sheet
            if "Datasheet" in oot_wb.sheetnames:
                ws_old = oot_wb["Datasheet"]
                oot_wb.remove(ws_old)
            oot_ds = oot_wb.create_sheet("Datasheet", index=0)
            for row in filtered_table:
                oot_ds.append(row)

            # -------- Parameter headers from PDF grey bars --------
            parameter_headers = extract_parameter_headers_from_pdf(ds_file)
            # Initialize ordered flags dict: all False by default
            for hdr in parameter_headers:
                parameter_flags[hdr] = False
            print(parameter_headers)
            print()
            print(parameter_flags)
            print()


            # -------- Map headers to row indices in Datasheet --------
            # We find each header row in order, scanning downward to avoid matching earlier rows
            scan_row = 1
            for hdr in parameter_headers:
                r = sheet_find_first_row_for_header(oot_ds, scan_row, hdr)
                if r is not None:
                    header_rows_in_sheet[hdr] = r
                    scan_row = r + 1  # continue scanning after this header
                else:
                    # If header text didn't land in the extracted table, skip it for regioning
                    # (it will simply never be marked True)
                    pass

            # -------- Build regions and set True if exact 'Fail' in column D --------
            ordered_hdrs_in_sheet = [h for h in parameter_headers if h in header_rows_in_sheet]
            for i, hdr in enumerate(ordered_hdrs_in_sheet):
                start_r = header_rows_in_sheet[hdr] + 1  # usually data starts below header row
                end_r = (header_rows_in_sheet[ordered_hdrs_in_sheet[i+1]] - 1) if i+1 < len(ordered_hdrs_in_sheet) else oot_ds.max_row
                if start_r <= end_r and region_has_exact_fail_col_d(oot_ds, start_r, end_r):
                    parameter_flags[hdr] = True

            # -------- Duplicate Impact Analysis for each True --------
            impact_sheet = oot_wb["Impact Analysis"]
            for hdr, flag in parameter_flags.items():
                if flag:
                    new_sheet = oot_wb.copy_worksheet(impact_sheet)
                    # Excel sheet name limit is 31 chars; also avoid illegal chars
                    safe_name = "".join(ch for ch in hdr if ch not in r'[]:*?/\\').strip()
                    new_sheet.title = (safe_name[:31] or "Impact Copy")

        else:
            print("Could not extract data from the datasheet PDF.")

# Default active sheet
oot_wb.active = oot_wb["Impact Analysis"]

# ---------------------------
# Save cleanly to XLSX (do NOT rename an XLSM)
# ---------------------------
final_xlsx = os.path.join(oot_dir, oot_out_file)
oot_wb.save(final_xlsx)
oot_wb.close()

print(f"Saved: {final_xlsx}")